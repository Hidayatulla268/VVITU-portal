"""
VVIT Portal — Faculty Views

Handles all faculty-facing functionality:
  • Dashboard overview
  • Mark / edit attendance (radio-button UI, AJAX-assisted)
  • View attendance reports with filters
  • Export reports to Excel (openpyxl) and PDF (reportlab)
  • Counselled-students list

HOD and Lab Technician roles share the same views through the middleware
permission mapping (both resolve to faculty:* URLs).
"""

import io
import logging
import datetime
from functools import wraps

logger = logging.getLogger(__name__)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Count, Q
from django.conf import settings
from django.views.decorators.http import require_POST

from accounts.models import Faculty, Student, Achievement, FacultyLeaveRequest
from core.models import (
    Section, Timetable, Attendance, Subject, Result, Exam, Year,
    FacultyAttendance, ClassTransfer, ClassDiary
)
from core.sms_utils import send_absent_notifications, send_absent_sms_to_parent


# ─────────────────────────────────────────────
# HELPER DECORATOR
# ─────────────────────────────────────────────
FACULTY_ROLES = {'faculty', 'hod', 'lab_technician'}

def faculty_required(view_func):
    """Ensures the user has a faculty-like role and a Faculty profile."""
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        if request.user.role not in FACULTY_ROLES:
            messages.error(request, "Access denied.")
            return redirect(request.user.get_dashboard_url())
        try:
            request.faculty = request.user.faculty_profile
        except Faculty.DoesNotExist:
            messages.error(request, "Faculty profile not found. Contact admin.")
            return redirect('accounts:login')
        return view_func(request, *args, **kwargs)
    return wrapper


# ─────────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────────
@faculty_required
def dashboard(request):
    """
    Summary dashboard:
      • Faculty's own attendance statistics (Days Present & Absent)
      • Timetable today with assigned classroom room numbers
      • Active class transfers / proxy classes today
      • Total sections and students handled
    """
    faculty  = request.faculty
    today    = timezone.localdate()
    day_name = today.strftime('%A')

    # Faculty Attendance statistics
    fac_att_qs = FacultyAttendance.objects.filter(faculty=faculty)
    faculty_present_days = fac_att_qs.filter(status='P').count()
    faculty_absent_days  = fac_att_qs.filter(status='A').count()
    faculty_leave_days   = fac_att_qs.filter(status='L').count()
    total_working_days   = fac_att_qs.count()
    faculty_att_pct      = round(faculty_present_days / total_working_days * 100, 1) if total_working_days > 0 else 100.0

    # Subjects and timetable entries for this faculty today (with room_number)
    timetable_today = (
        Timetable.objects
        .filter(faculty=faculty, day=day_name)
        .select_related('section', 'subject', 'section__branch')
        .order_by('period')
    )

    # Full weekly timetable with room numbers
    weekly_timetable = (
        Timetable.objects
        .filter(faculty=faculty)
        .select_related('section', 'subject', 'section__branch')
        .order_by('day', 'period')
    )

    subjects = (
        Subject.objects
        .filter(faculty=faculty, is_deleted=False)
        .select_related('branch', 'year')
    )

    # Sections this faculty handles (via timetable)
    sections = (
        Timetable.objects
        .filter(faculty=faculty)
        .values_list('section', flat=True)
        .distinct()
    )
    section_count = sections.count()
    student_count = Student.objects.filter(section__in=sections, user__is_deleted=False).count()

    # Transferred classes for today
    # 1. Received (Proxy assigned to logged-in faculty)
    transferred_today = (
        ClassTransfer.objects
        .filter(substitute_faculty=faculty, date=today)
        .select_related('timetable_entry__section', 'timetable_entry__subject', 'original_faculty__user')
    )

    # 2. Transferred Out (Given to someone else today)
    transferred_given_today = (
        ClassTransfer.objects
        .filter(original_faculty=faculty, date=today)
        .select_related('timetable_entry__section', 'timetable_entry__subject', 'substitute_faculty__user')
    )

    # Department faculty for class transfer dropdown
    department_faculty = Faculty.objects.filter(is_active=True).exclude(id=faculty.id).select_related('user', 'department')
    if faculty.department:
        department_faculty = department_faculty.filter(department=faculty.department)

    context = {
        'faculty':              faculty,
        'faculty_present_days': faculty_present_days,
        'faculty_absent_days':  faculty_absent_days,
        'faculty_leave_days':   faculty_leave_days,
        'total_working_days':   total_working_days,
        'faculty_att_pct':      faculty_att_pct,
        'timetable_today':      timetable_today,
        'weekly_timetable':     weekly_timetable,
        'subjects':             subjects,
        'section_count':        section_count,
        'student_count':        student_count,
        'transferred_today':    transferred_today,
        'transferred_given_today': transferred_given_today,
        'department_faculty':   department_faculty,
        'today':                today,
    }
    return render(request, 'faculty/dashboard.html', context)


# ─────────────────────────────────────────────
# AJAX — Load students for a section
# ─────────────────────────────────────────────
@faculty_required
def ajax_get_students(request):
    """
    Return JSON list of students in a section (for attendance form).
    If date and slot_id are provided, pre-populate each student's saved attendance status ('P' or 'A')
    and any saved ClassDiary lesson details for that slot and date.
    """
    section_id = request.GET.get('section_id')
    date_str   = request.GET.get('date')
    slot_id    = request.GET.get('slot_id')

    if not section_id:
        return JsonResponse({'error': 'section_id required'}, status=400)

    students = (
        Student.objects
        .filter(section_id=section_id, is_active=True, user__is_deleted=False)
        .select_related('user')
        .order_by('roll_number')
        .only('id', 'roll_number', 'user__first_name', 'user__last_name')
    )

    # Check for existing attendance records & diary notes
    att_map = {}
    already_marked = False
    diary_data = None

    if slot_id and date_str:
        try:
            att_date = datetime.date.fromisoformat(date_str)
            records = Attendance.objects.filter(
                timetable_entry_id=slot_id,
                date=att_date,
                student__section_id=section_id
            ).values('student_id', 'status')

            for r in records:
                att_map[r['student_id']] = r['status']

            if att_map:
                already_marked = True

            # Check if ClassDiary was previously saved for this slot + date
            diary = ClassDiary.objects.filter(timetable_entry_id=slot_id, date=att_date).first()
            if diary:
                diary_data = {
                    'unit_number': diary.unit_number,
                    'topic_covered': diary.topic_covered or '',
                    'discussion_summary': diary.discussion_summary or '',
                    'homework_assignment': diary.homework_assignment or '',
                }
        except (ValueError, TypeError):
            pass

    data = [
        {
            'id':          s.id,
            'roll_number': s.roll_number,
            'name':        s.user.get_full_name(),
            'status':      att_map.get(s.id, 'P'),  # Default to 'P' if not marked before
            'is_saved':    s.id in att_map,
        }
        for s in students
    ]

    return JsonResponse({
        'students':       data,
        'already_marked': already_marked,
        'marked_count':   len(att_map),
        'total_count':    len(data),
        'diary':          diary_data,
    })


# ─────────────────────────────────────────────
# AJAX — Load timetable slots for section + day
# ─────────────────────────────────────────────
@faculty_required
def ajax_get_timetable(request):
    """Return JSON timetable slots for (section, day) with timings, locations, and auto-mapping."""
    section_id = request.GET.get('section_id')
    day        = request.GET.get('day')
    faculty    = request.faculty
    if not section_id or not day:
        return JsonResponse({'error': 'section_id and day required'}, status=400)

    slots = (
        Timetable.objects
        .filter(section_id=section_id, day=day)
        .select_related('subject', 'faculty__user')
        .order_by('period')
    )
    period_timings = {
        1: "09:00 AM - 09:50 AM",
        2: "09:50 AM - 10:40 AM",
        3: "10:50 AM - 11:40 AM",
        4: "11:40 AM - 12:30 PM",
        5: "01:20 PM - 02:10 PM",
        6: "02:10 PM - 03:00 PM",
        7: "03:10 PM - 04:00 PM",
        8: "04:00 PM - 04:50 PM",
    }
    data = []
    for s in slots:
        start_t = s.start_time.strftime("%I:%M %p") if getattr(s, 'start_time', None) else None
        end_t   = s.end_time.strftime("%I:%M %p") if getattr(s, 'end_time', None) else None
        timing_str = f"{start_t} - {end_t}" if (start_t and end_t) else period_timings.get(s.period, "Scheduled Slot")
        is_my_slot = (s.faculty == faculty)
        data.append({
            'id':           s.id,
            'period':       s.period,
            'subject_code':  s.subject.code,
            'subject_name':  s.subject.name,
            'subject_short': s.subject.short_name,
            'room_number':  getattr(s, 'room_number', 'Room 101') or 'Room 101',
            'timing':       timing_str,
            'faculty_name': s.faculty.full_name if s.faculty else "Faculty",
            'is_my_slot':   is_my_slot,
        })
    return JsonResponse({'slots': data})


@login_required
def ajax_get_free_faculty(request):
    """Return JSON list of free faculty for a specific (date, period)."""
    date_str = request.GET.get('date')
    period_str = request.GET.get('period')
    branch_id_str = request.GET.get('branch_id') or request.GET.get('department')
    exclude_fac_id = request.GET.get('exclude_faculty_id')

    if not date_str or not period_str:
        return JsonResponse({'error': 'date and period parameters required'}, status=400)

    from core.transfer_utils import get_free_faculty_for_period, parse_flexible_date
    from core.models import Branch
    from accounts.models import Faculty

    req_date = parse_flexible_date(date_str)
    if not req_date:
        return JsonResponse({'error': 'Invalid date format'}, status=400)

    try:
        period = int(period_str)
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid period format'}, status=400)


    current_fac = getattr(request.user, 'faculty_profile', None)

    dept = None
    if branch_id_str and branch_id_str.isdigit():
        dept = Branch.objects.filter(id=int(branch_id_str)).first()
    elif request.user.role != 'admin' and current_fac:
        dept = current_fac.department

    exclude_fac = None
    if exclude_fac_id and exclude_fac_id.isdigit():
        exclude_fac = Faculty.objects.filter(id=int(exclude_fac_id)).first()
    elif current_fac:
        exclude_fac = current_fac

    free_fac_qs = get_free_faculty_for_period(
        date=req_date,
        period=period,
        department=dept,
        exclude_faculty=exclude_fac
    )

    data = [
        {
            'id': f.id,
            'name': f.full_name,
            'employee_id': f.employee_id,
            'department': f.department.code if f.department else 'Gen',
        }
        for f in free_fac_qs
    ]
    return JsonResponse({'free_faculty': data})



# ─────────────────────────────────────────────
# MARK ATTENDANCE
# ─────────────────────────────────────────────
@faculty_required
def mark_attendance(request):
    """
    Two-phase view:
      GET  — render the filter form (course, branch, section, date, timetable slot).
      POST — save attendance records & send SMS to parents if absent.
    """
    faculty    = request.faculty
    edit_window = getattr(settings, 'ATTENDANCE_EDIT_WINDOW_DAYS', 2)
    today       = timezone.localdate()
    min_date    = today - datetime.timedelta(days=edit_window - 1)

    # Sections where this faculty teaches or has proxy transfer today
    section_ids = list(
        Timetable.objects
        .filter(faculty=faculty)
        .values_list('section_id', flat=True)
        .distinct()
    )
    proxy_sec_ids = list(
        ClassTransfer.objects
        .filter(substitute_faculty=faculty, date=today)
        .values_list('timetable_entry__section_id', flat=True)
    )
    all_sec_ids = list(set(section_ids + proxy_sec_ids))
    sections = Section.objects.filter(id__in=all_sec_ids).select_related('branch', 'year')

    if request.method == 'POST':
        section_id  = request.POST.get('section')
        date_str    = request.POST.get('date')
        slot_id     = request.POST.get('slot')

        try:
            att_date = datetime.date.fromisoformat(date_str)
        except (ValueError, TypeError):
            messages.error(request, "Invalid date format.")
            return redirect('faculty:mark_attendance')

        # Enforce edit window
        if att_date < min_date or att_date > today:
            messages.error(
                request,
                f"Attendance can only be marked for {today.strftime('%d %b')} "
                f"back to {min_date.strftime('%d %b %Y')}."
            )
            return redirect('faculty:mark_attendance')

        slot = get_object_or_404(Timetable, id=slot_id)

        # Check if faculty is original faculty on approved leave for att_date
        from accounts.models import FacultyLeaveRequest
        if slot.faculty == faculty:
            on_leave = FacultyLeaveRequest.objects.filter(
                faculty=faculty,
                start_date__lte=att_date,
                end_date__gte=att_date,
                status='approved'
            ).exists()
            if on_leave:
                messages.error(
                    request,
                    f"You are on approved leave for {att_date.strftime('%d-%b-%Y')}. You cannot post attendance for your classes. Substitute faculty must post attendance."
                )
                return redirect('faculty:mark_attendance')

        students_in_section = Student.objects.filter(section_id=section_id, is_active=True, user__is_deleted=False)

        saved_count = 0
        sms_count = 0
        for student in students_in_section:
            field_name = f"attendance_{student.id}"
            status     = request.POST.get(field_name, 'A')
            if status not in ('P', 'A'):
                status = 'A'
            
            att_rec, created = Attendance.objects.update_or_create(
                student=student,
                timetable_entry=slot,
                date=att_date,
                defaults={'status': status, 'marked_by': faculty},
            )
            saved_count += 1

            # Trigger absent notifications to parent & student if marked absent
            if status == 'A':
                from core.sms_utils import send_absent_notifications
                sent = send_absent_notifications(student, slot, att_date)
                if sent:
                    sms_count += 1

        # Mark ClassTransfer as completed if substitute faculty marked attendance
        ClassTransfer.objects.filter(timetable_entry=slot, date=att_date).update(status='completed')

        # Optional Class Discussion / Lesson Log
        topic_covered = request.POST.get('topic_covered', '').strip()
        if topic_covered:
            unit_val = request.POST.get('unit_number', '').strip()
            unit_number = int(unit_val) if unit_val.isdigit() and 1 <= int(unit_val) <= 6 else 1
            # Auto-detect unit from topic string if default 1 was used and text mentions Unit 2..5
            if unit_number == 1:
                import re
                m = re.search(r'unit\s*([1-5])', topic_covered, re.IGNORECASE)
                if m:
                    unit_number = int(m.group(1))

            discussion_summary = request.POST.get('discussion_summary', '').strip()
            homework_assignment = request.POST.get('homework_assignment', '').strip()
            ClassDiary.objects.update_or_create(
                timetable_entry=slot,
                date=att_date,
                defaults={
                    'section': slot.section,
                    'subject': slot.subject,
                    'faculty': faculty,
                    'period': slot.period,
                    'unit_number': unit_number,
                    'topic_covered': topic_covered,
                    'discussion_summary': discussion_summary,
                    'homework_assignment': homework_assignment,
                }
            )

        msg_text = f"Attendance saved for {saved_count} students."
        if topic_covered:
            msg_text += " Class discussion notes recorded."
        if sms_count > 0:
            msg_text += f" ({sms_count} absence SMS notification(s) sent to parents)."
        messages.success(request, msg_text)
        return redirect('faculty:mark_attendance')

    context = {
        'sections':    sections,
        'today':       today.isoformat(),
        'min_date':    min_date.isoformat(),
        'faculty':     faculty,
        'unit_choices': ClassDiary.UNIT_CHOICES,
    }
    return render(request, 'faculty/mark_attendance.html', context)


# ─────────────────────────────────────────────
# CLASS DIARY / LESSON DISCUSSION LOGS
# ─────────────────────────────────────────────
@faculty_required
def class_diary(request):
    """
    Dedicated view for faculty to view, add, edit, and search daily class discussion logs.
    """
    faculty = request.faculty
    today = timezone.localdate()
    from core.transfer_utils import parse_flexible_date

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'save_entry':
            entry_id = request.POST.get('entry_id')
            slot_id = request.POST.get('slot_id')
            date_str = request.POST.get('date', '').strip()
            unit_val = request.POST.get('unit_number', '').strip()
            topic_covered = request.POST.get('topic_covered', '').strip()
            discussion_summary = request.POST.get('discussion_summary', '').strip()
            homework_assignment = request.POST.get('homework_assignment', '').strip()

            entry_date = parse_flexible_date(date_str) or today
            unit_number = int(unit_val) if unit_val.isdigit() and 1 <= int(unit_val) <= 6 else 1

            if not topic_covered:
                messages.error(request, "Topic covered is required.")
                return redirect('faculty:class_diary')

            if entry_id and entry_id.isdigit():
                entry = get_object_or_404(ClassDiary, id=int(entry_id), faculty=faculty)
                entry.unit_number = unit_number
                entry.topic_covered = topic_covered
                entry.discussion_summary = discussion_summary
                entry.homework_assignment = homework_assignment
                entry.save()
                messages.success(request, f"Class log for {entry.subject.code} on {entry.date.strftime('%d-%b-%Y')} updated successfully.")
            elif slot_id and slot_id.isdigit():
                slot = get_object_or_404(Timetable, id=int(slot_id))
                ClassDiary.objects.update_or_create(
                    timetable_entry=slot,
                    date=entry_date,
                    defaults={
                        'section': slot.section,
                        'subject': slot.subject,
                        'faculty': faculty,
                        'period': slot.period,
                        'unit_number': unit_number,
                        'topic_covered': topic_covered,
                        'discussion_summary': discussion_summary,
                        'homework_assignment': homework_assignment,
                    }
                )
                messages.success(request, f"Class discussion log for {slot.subject.code} ({slot.section}) recorded successfully.")
            return redirect('faculty:class_diary')

        elif action == 'delete_entry':
            entry_id = request.POST.get('entry_id')
            if entry_id and entry_id.isdigit():
                entry = get_object_or_404(ClassDiary, id=int(entry_id), faculty=faculty)
                subj_code = entry.subject.code
                d_str = entry.date.strftime('%d-%b-%Y')
                entry.delete()
                messages.success(request, f"Class log for {subj_code} on {d_str} deleted.")
            return redirect('faculty:class_diary')

    # Filters
    search_query = request.GET.get('search', '').strip()
    section_id = request.GET.get('section_id', '').strip()
    subject_id = request.GET.get('subject_id', '').strip()
    unit_filter = request.GET.get('unit_number', '').strip()
    date_from_str = request.GET.get('date_from', '').strip()
    date_to_str = request.GET.get('date_to', '').strip()

    diary_qs = ClassDiary.objects.filter(faculty=faculty).select_related('section__branch', 'subject', 'timetable_entry')

    if search_query:
        diary_qs = diary_qs.filter(
            Q(topic_covered__icontains=search_query) |
            Q(discussion_summary__icontains=search_query) |
            Q(homework_assignment__icontains=search_query) |
            Q(subject__name__icontains=search_query) |
            Q(subject__code__icontains=search_query)
        )

    if section_id and section_id.isdigit():
        diary_qs = diary_qs.filter(section_id=int(section_id))

    if subject_id and subject_id.isdigit():
        diary_qs = diary_qs.filter(subject_id=int(subject_id))

    if unit_filter and unit_filter.isdigit():
        diary_qs = diary_qs.filter(unit_number=int(unit_filter))

    date_from = parse_flexible_date(date_from_str)
    date_to = parse_flexible_date(date_to_str)

    if date_from:
        diary_qs = diary_qs.filter(date__gte=date_from)
    if date_to:
        diary_qs = diary_qs.filter(date__lte=date_to)

    entries = diary_qs.order_by('-date', 'period')

    # Sections & Subjects handled by this faculty
    handled_sec_ids = Timetable.objects.filter(faculty=faculty).values_list('section_id', flat=True).distinct()
    handled_sections = Section.objects.filter(id__in=handled_sec_ids).select_related('branch', 'year')
    handled_subjects = Subject.objects.filter(faculty=faculty, is_deleted=False).select_related('branch', 'year')

    # Today's slots for quick modal selection
    day_name = today.strftime('%A')
    today_slots = Timetable.objects.filter(faculty=faculty, day__iexact=day_name).select_related('section__branch', 'subject').order_by('period')

    # All weekly slots for faculty
    all_slots = Timetable.objects.filter(faculty=faculty).select_related('section__branch', 'subject').order_by('day', 'period')

    context = {
        'entries': entries,
        'handled_sections': handled_sections,
        'handled_subjects': handled_subjects,
        'today_slots': today_slots,
        'all_slots': all_slots,
        'search_query': search_query,
        'section_id': int(section_id) if section_id and section_id.isdigit() else '',
        'subject_id': int(subject_id) if subject_id and subject_id.isdigit() else '',
        'unit_number': int(unit_filter) if unit_filter and unit_filter.isdigit() else '',
        'unit_choices': ClassDiary.UNIT_CHOICES,
        'date_from': date_from_str,
        'date_to': date_to_str,
        'today': today,
    }
    return render(request, 'faculty/class_diary.html', context)


# ─────────────────────────────────────────────
# CLASS TRANSFER / PROXY ACTION
# ─────────────────────────────────────────────
@faculty_required
def transfer_class(request):
    """
    Allows faculty to transfer a timetable class period for today (or selected date)
    to another available faculty member as a proxy/substitute.
    """
    faculty = request.faculty
    today   = timezone.localdate()

    if request.method == 'POST':
        timetable_id = request.POST.get('timetable_id')
        substitute_id = request.POST.get('substitute_id')
        reason = request.POST.get('reason', '').strip()
        date_str = request.POST.get('date', today.isoformat())

        try:
            transfer_date = datetime.date.fromisoformat(date_str)
        except (ValueError, TypeError):
            transfer_date = today

        slot = get_object_or_404(Timetable, id=timetable_id, faculty=faculty)
        substitute = get_object_or_404(Faculty, id=substitute_id, is_active=True)

        if substitute == faculty:
            messages.error(request, "Cannot transfer class to yourself.")
            return redirect('faculty:dashboard')

        # Verify substitute faculty is FREE during this period
        from core.transfer_utils import get_free_faculty_for_period
        from core.sms_utils import send_class_transfer_notification

        free_fac = get_free_faculty_for_period(
            date=transfer_date,
            period=slot.period,
            department=faculty.department,
            exclude_faculty=faculty
        )

        if substitute not in free_fac:
            messages.error(
                request,
                f"Prof. {substitute.full_name} has another scheduled class or proxy assignment during Period {slot.period} on {transfer_date.strftime('%d-%b-%Y')}. Please select an available free faculty member."
            )
            return redirect('faculty:dashboard')

        transfer, created = ClassTransfer.objects.update_or_create(
            timetable_entry=slot,
            date=transfer_date,
            defaults={
                'original_faculty': faculty,
                'substitute_faculty': substitute,
                'reason': reason or 'Faculty absent / on leave',
                'status': 'accepted',
            }
        )

        # Send instant SMS & Email notification to substitute faculty
        send_class_transfer_notification(transfer)

        messages.success(
            request,
            f"Class P{slot.period} ({slot.subject.code}) transferred to Prof. {substitute.full_name} for {transfer_date.strftime('%d-%b-%Y')}. SMS & Email notification dispatched."
        )
        return redirect('faculty:dashboard')

    return redirect('faculty:dashboard')


# ─────────────────────────────────────────────
# FACULTY OWN ATTENDANCE REPORT
# ─────────────────────────────────────────────
@faculty_required
def my_attendance(request):
    """
    Displays logged-in faculty member's attendance history.
    Supports Month-wise filtering (e.g. 2026-08) and Custom Date Range filtering.
    """
    faculty = request.faculty
    today   = timezone.localdate()

    month_year = request.GET.get('month_year', '')
    date_from  = request.GET.get('date_from', '')
    date_to    = request.GET.get('date_to', '')

    qs = FacultyAttendance.objects.filter(faculty=faculty).select_related('marked_by')

    if month_year:
        try:
            yr, mn = map(int, month_year.split('-'))
            qs = qs.filter(date__year=yr, date__month=mn)
        except ValueError:
            pass
    elif date_from or date_to:
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
    else:
        # Default to current month
        qs = qs.filter(date__year=today.year, date__month=today.month)
        month_year = today.strftime('%Y-%m')

    records = qs.order_by('-date')

    present_count = records.filter(status='P').count()
    absent_count  = records.filter(status='A').count()
    leave_count   = records.filter(status='L').count()
    total_days    = records.count()
    percentage    = round(present_count / total_days * 100, 1) if total_days > 0 else 0.0

    context = {
        'faculty':       faculty,
        'records':       records,
        'present_count': present_count,
        'absent_count':  absent_count,
        'leave_count':   leave_count,
        'total_days':    total_days,
        'percentage':    percentage,
        'month_year':    month_year,
        'date_from':     date_from,
        'date_to':       date_to,
    }
    return render(request, 'faculty/my_attendance.html', context)


# ─────────────────────────────────────────────
# ATTENDANCE REPORTS (STUDENT ATTENDANCE)
# ─────────────────────────────────────────────
@faculty_required
def reports(request):
    """
    Display attendance report filtered by section, subject, month-wise or custom date range.
    Shows student-wise totals and percentages in a table.
    """
    faculty  = request.faculty
    today    = timezone.localdate()

    section_ids = (
        Timetable.objects
        .filter(faculty=faculty)
        .values_list('section_id', flat=True)
        .distinct()
    )
    sections = Section.objects.filter(id__in=section_ids).select_related('branch', 'year')
    subjects = Subject.objects.filter(faculty=faculty, is_deleted=False).select_related('branch', 'year')

    # Read filter params
    section_id = request.GET.get('section')
    subject_id = request.GET.get('subject')
    month_year = request.GET.get('month_year', '')
    date_from  = request.GET.get('date_from', '')
    date_to    = request.GET.get('date_to', '')

    if month_year:
        try:
            yr, mn = map(int, month_year.split('-'))
            import calendar
            last_day = calendar.monthrange(yr, mn)[1]
            date_from = f"{yr:04d}-{mn:02d}-01"
            date_to   = f"{yr:04d}-{mn:02d}-{last_day:02d}"
        except ValueError:
            pass

    if not date_from:
        date_from = (today - datetime.timedelta(days=30)).isoformat()
    if not date_to:
        date_to = today.isoformat()

    report_data = []
    if section_id:
        att_qs = (
            Attendance.objects
            .filter(
                timetable_entry__section_id=section_id,
                date__gte=date_from,
                date__lte=date_to,
            )
            .select_related('student__user', 'timetable_entry__subject')
        )
        if subject_id:
            att_qs = att_qs.filter(timetable_entry__subject_id=subject_id)

        # Aggregate per student — pre-populate with all section students
        section_students = Student.objects.filter(section_id=section_id, is_active=True, user__is_deleted=False).select_related('user')
        student_map = {
            st.id: {
                'student_id': st.id,
                'roll':    st.roll_number,
                'name':    st.user.get_full_name(),
                'total':   0,
                'present': 0,
            }
            for st in section_students
        }
        for rec in att_qs:
            sid = rec.student.id
            if sid not in student_map:
                student_map[sid] = {
                    'student_id': sid,
                    'roll':    rec.student.roll_number,
                    'name':    rec.student.user.get_full_name(),
                    'total':   0,
                    'present': 0,
                }
            student_map[sid]['total'] += 1
            if rec.status == 'P':
                student_map[sid]['present'] += 1

        for sid, d in student_map.items():
            t = d['total']
            p = d['present']
            d['pct']    = round(p / t * 100, 1) if t else 0
            d['absent'] = t - p
            d['low']    = d['pct'] < getattr(settings, 'LOW_ATTENDANCE_THRESHOLD', 75)
            report_data.append(d)

        report_data.sort(key=lambda x: x['roll'])

    context = {
        'sections':    sections,
        'subjects':    subjects,
        'report_data': report_data,
        'section_id':  section_id,
        'subject_id':  subject_id,
        'month_year':  month_year,
        'date_from':   date_from,
        'date_to':     date_to,
        'threshold':   getattr(settings, 'LOW_ATTENDANCE_THRESHOLD', 75),
    }
    return render(request, 'faculty/reports.html', context)


# ─────────────────────────────────────────────
# EXPORT — EXCEL
# ─────────────────────────────────────────────
@faculty_required
def export_excel(request):
    """Export the current attendance report as an .xlsx file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messages.error(request, "openpyxl is not installed. Cannot export Excel.")
        return redirect('faculty:reports')

    faculty    = request.faculty
    section_id = request.GET.get('section')
    subject_id = request.GET.get('subject')
    date_from  = request.GET.get('date_from')
    date_to    = request.GET.get('date_to')
    today      = timezone.localdate()

    att_qs = (
        Attendance.objects
        .filter(
            timetable_entry__section_id=section_id or None,
            date__gte=date_from or (today - datetime.timedelta(days=30)).isoformat(),
            date__lte=date_to   or today.isoformat(),
        )
        .select_related('student__user', 'timetable_entry__subject')
    )
    if subject_id:
        att_qs = att_qs.filter(timetable_entry__subject_id=subject_id)

    # Aggregate
    student_map = {}
    for rec in att_qs:
        sid = rec.student.id
        if sid not in student_map:
            student_map[sid] = {
                'roll':    rec.student.roll_number,
                'name':    rec.student.user.get_full_name(),
                'total':   0,
                'present': 0,
            }
        student_map[sid]['total'] += 1
        if rec.status == 'P':
            student_map[sid]['present'] += 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance Report'

    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='CC0000')
    center      = Alignment(horizontal='center')

    headers = ['S.No', 'Roll Number', 'Student Name', 'Total Classes', 'Present', 'Absent', 'Percentage']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = center

    for row_idx, (_, d) in enumerate(student_map.items(), 2):
        t   = d['total']
        p   = d['present']
        pct = round(p / t * 100, 1) if t else 0
        ws.append([row_idx - 1, d['roll'], d['name'], t, p, t - p, pct])

    ws.column_dimensions['C'].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="attendance_report.xlsx"'
    return response


# ─────────────────────────────────────────────
# EXPORT — PDF
# ─────────────────────────────────────────────
@faculty_required
def export_pdf(request):
    """Export attendance report as a PDF using reportlab."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib         import colors
        from reportlab.platypus    import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles  import getSampleStyleSheet
    except ImportError:
        messages.error(request, "reportlab is not installed. Cannot export PDF.")
        return redirect('faculty:reports')

    faculty    = request.faculty
    section_id = request.GET.get('section')
    subject_id = request.GET.get('subject')
    date_from  = request.GET.get('date_from')
    date_to    = request.GET.get('date_to')
    today      = timezone.localdate()

    att_qs = (
        Attendance.objects
        .filter(
            timetable_entry__section_id=section_id or None,
            date__gte=date_from or (today - datetime.timedelta(days=30)).isoformat(),
            date__lte=date_to   or today.isoformat(),
        )
        .select_related('student__user', 'timetable_entry__subject')
    )
    if subject_id:
        att_qs = att_qs.filter(timetable_entry__subject_id=subject_id)

    student_map = {}
    for rec in att_qs:
        sid = rec.student.id
        if sid not in student_map:
            student_map[sid] = {
                'roll':    rec.student.roll_number,
                'name':    rec.student.user.get_full_name(),
                'total':   0,
                'present': 0,
            }
        student_map[sid]['total'] += 1
        if rec.status == 'P':
            student_map[sid]['present'] += 1

    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elems  = []

    # Title
    elems.append(Paragraph("VVIT — Attendance Report", styles['Title']))
    elems.append(Paragraph(f"Generated: {today.strftime('%d %B %Y')}", styles['Normal']))
    elems.append(Spacer(1, 12))

    # Table
    table_data = [['S.No', 'Roll Number', 'Student Name', 'Total', 'Present', 'Absent', 'Percentage']]
    for i, (_, d) in enumerate(student_map.items(), 1):
        t   = d['total']
        p   = d['present']
        pct = f"{round(p / t * 100, 1) if t else 0}%"
        table_data.append([i, d['roll'], d['name'], t, p, t - p, pct])

    tbl = Table(table_data, colWidths=[40, 90, 180, 60, 60, 60, 70])
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#CC0000')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('GRID',       (0,0), (-1,-1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#FFF0F0')]),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
        ('ALIGN',      (2,0), (2,-1), 'LEFT'),
    ]))
    elems.append(tbl)
    doc.build(elems)

    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="attendance_report.pdf"'
    return response


# ─────────────────────────────────────────────
# COUNSELLED & ASSIGNED STUDENTS
# ─────────────────────────────────────────────
@faculty_required
def counselled_students(request):
    """
    List all students associated with this faculty member:
      1. Counselled students (designated counsellor)
      2. Class students (designated class teacher)
      3. Subject students (students in sections taught by this faculty)
    """
    faculty = request.faculty
    
    # 1. Counselled Students
    counselled_list = (
        Student.objects
        .filter(counsellor=faculty, is_active=True, user__is_deleted=False)
        .select_related('user', 'branch', 'year', 'section')
        .order_by('roll_number')
    )
    
    # 2. Class Students
    class_list = (
        Student.objects
        .filter(class_teacher=faculty, is_active=True, user__is_deleted=False)
        .select_related('user', 'branch', 'year', 'section')
        .order_by('roll_number')
    )
    
    # 3. Subject Students (students in sections handled by this faculty)
    timetable_slots = Timetable.objects.filter(faculty=faculty).select_related('section', 'subject')
    section_ids = list(timetable_slots.values_list('section_id', flat=True).distinct())
    
    # Mapping of section ID to subjects taught
    section_subjects = {}
    for slot in timetable_slots:
        if slot.section_id not in section_subjects:
            section_subjects[slot.section_id] = []
        if slot.subject.code not in section_subjects[slot.section_id]:
            section_subjects[slot.section_id].append(slot.subject.code)
            
    subject_list = (
        Student.objects
        .filter(section_id__in=section_ids, is_active=True, user__is_deleted=False)
        .select_related('user', 'branch', 'year', 'section')
        .order_by('section__name', 'roll_number')
    )
    
    # Attach subject codes to each subject student for rendering
    for student in subject_list:
        student.subjects_taught = ", ".join(section_subjects.get(student.section_id, []))

    context = {
        'faculty': faculty,
        'counselled_students': counselled_list,
        'class_students': class_list,
        'subject_students': subject_list,
    }
    return render(request, 'faculty/counselled_students.html', context)


# ─────────────────────────────────────────────
# STUDENT RESULTS
# ─────────────────────────────────────────────
@faculty_required
def student_results(request):
    """
    View for class teachers and counsellors to see their assigned students' exam results.
    """
    faculty = request.faculty
    
    # Get students where faculty is class_teacher or counsellor (HOD sees all students in branch)
    if request.user.role == 'hod':
        students = Student.objects.filter(
            branch=faculty.department,
            is_active=True,
            user__is_deleted=False
        ).select_related('user')
    else:
        students = Student.objects.filter(
            Q(class_teacher=faculty) | Q(counsellor=faculty),
            is_active=True,
            user__is_deleted=False
        ).select_related('user')
    
    selected_student_id = request.GET.get('student', '')
    selected_exam_id = request.GET.get('exam', '')
    
    selected_student = None
    selected_exam_obj = None
    results_list = []
    sgpa = 0.0
    pass_status = "Pass"
    revaluation_date = None
    
    # List of all exams for branch/year combination of these students
    branch_ids = students.values_list('branch_id', flat=True).distinct()
    year_ids = students.values_list('year_id', flat=True).distinct()
    exams = Exam.objects.filter(branch_id__in=branch_ids, year_id__in=year_ids).order_by('-date')
    
    if selected_student_id and selected_exam_id:
        try:
            selected_student = students.get(id=selected_student_id)
            selected_exam_obj = Exam.objects.get(id=selected_exam_id)
            
            results_qs = Result.objects.filter(
                student=selected_student,
                exam_id=selected_exam_id,
                exam__release__released=True
            ).select_related('subject').order_by('subject__name')
            
            results_list = list(results_qs)
            
            # Calculate SGPA and Pass/Fail status dynamically according to R23 regulation
            grade_points = {
                'S': 10, 'A': 9, 'B': 8, 'C': 7, 'D': 6, 'E': 5,
                'F': 0, 'Ab': 0
            }
            total_points = 0
            total_credits = 0
            has_fail = False
            
            for r in results_list:
                g = r.grade
                if g in ['CP', 'NCP']:
                    if g == 'NCP':
                        has_fail = True
                    continue
                if g in ['F', 'Ab'] or not g:
                    has_fail = True
                
                credits = r.subject.credits if r.subject else 3
                points = grade_points.get(g, 0)
                total_points += points * credits
                total_credits += credits
                
            sgpa = round(total_points / total_credits, 2) if total_credits > 0 else 0.0
            pass_status = "Fail" if has_fail else "Pass"
            
            if selected_exam_obj.date:
                revaluation_date = selected_exam_obj.date + datetime.timedelta(days=40)
            else:
                revaluation_date = timezone.localdate() + datetime.timedelta(days=30)
                
        except (Student.DoesNotExist, Exam.DoesNotExist):
            pass
            
    # Default view list
    results = Result.objects.filter(
        student__in=students,
        exam__release__released=True
    ).select_related('student__user', 'exam', 'subject').order_by('student__roll_number', '-exam__date')
    
    if selected_student_id:
        results = results.filter(student_id=selected_student_id)
    if selected_exam_id:
        results = results.filter(exam_id=selected_exam_id)
        
    context = {
        'faculty':             faculty,
        'students':            students,
        'exams':               exams,
        'selected_student':    selected_student,
        'selected_exam_obj':   selected_exam_obj,
        'selected_student_id': selected_student_id,
        'selected_exam_id':    selected_exam_id,
        'results_list':        results_list,
        'sgpa':                sgpa,
        'pass_status':         pass_status,
        'revaluation_date':    revaluation_date,
        'results':             results,
    }
    return render(request, 'faculty/student_results.html', context)


# ─────────────────────────────────────────────
# FACULTY UPLOAD MARKS
# ─────────────────────────────────────────────
@faculty_required
def upload_marks(request):
    """
    Allows faculty to input or upload marks for a specific subject, exam, and section (class).
    """
    import csv
    import io
    from django.db import transaction
    
    faculty = request.faculty
    years = Year.objects.all()
    year_id = request.GET.get('year') or request.POST.get('year')

    if request.user.role == 'hod':
        subjects = Subject.objects.filter(branch=faculty.department, is_deleted=False).select_related('branch', 'year')
    else:
        subjects = Subject.objects.filter(faculty=faculty, is_deleted=False).select_related('branch', 'year')

    selected_subject_id = request.GET.get('subject', '')
    selected_exam_id = request.GET.get('exam', '')
    selected_section_id = request.GET.get('section', '')

    # Validate parameters based on selected year
    if year_id:
        if selected_subject_id and not subjects.filter(id=selected_subject_id, year_id=year_id).exists():
            selected_subject_id = ''
        if selected_exam_id and not Exam.objects.filter(id=selected_exam_id, year_id=year_id).exists():
            selected_exam_id = ''
        if selected_section_id and not Section.objects.filter(id=selected_section_id, year_id=year_id).exists():
            selected_section_id = ''
    else:
        selected_subject_id = ''
        selected_exam_id = ''
        selected_section_id = ''

    # Apply year filtering
    if year_id:
        subjects = subjects.filter(year_id=year_id)
        branch_ids = subjects.values_list('branch_id', flat=True).distinct()
        exams = Exam.objects.filter(branch_id__in=branch_ids, year_id=year_id).exclude(exam_type='final').order_by('-date')
    else:
        subjects = Subject.objects.none()
        exams = Exam.objects.none()

    selected_subject = None
    selected_exam = None
    selected_section = None
    sections = []
    students = []
    current_results = {}
    
    if selected_subject_id:
        if request.user.role == 'hod':
            selected_subject = get_object_or_404(Subject, id=selected_subject_id, branch=faculty.department)
            sections = Section.objects.filter(branch=faculty.department, year=selected_subject.year).distinct()
        else:
            selected_subject = get_object_or_404(Subject, id=selected_subject_id, faculty=faculty)
            sections = Section.objects.filter(timetable_entries__subject=selected_subject).distinct()
        
    if selected_exam_id:
        selected_exam = get_object_or_404(Exam, id=selected_exam_id)
        if selected_exam.exam_type == 'final':
            messages.error(request, "Only the Administrator is authorized to upload Semester Final exam results.")
            return redirect('faculty:upload_marks')
        
    if selected_section_id and selected_subject and selected_exam:
        if request.user.role == 'hod':
            selected_section = get_object_or_404(Section, id=selected_section_id, branch=faculty.department)
        else:
            selected_section = get_object_or_404(Section, id=selected_section_id)
        students = Student.objects.filter(section=selected_section, is_active=True, user__is_deleted=False).select_related('user').order_by('roll_number')
        
        # Load existing results for these students, exam, and subject
        results_qs = Result.objects.filter(
            student__in=students,
            exam=selected_exam,
            subject=selected_subject
        )
        current_results = {r.student_id: r for r in results_qs}
        
    if request.method == 'POST':
        subj_id = request.POST.get('subject')
        ex_id = request.POST.get('exam')
        sec_id = request.POST.get('section')
        
        # Resolve objects
        if request.user.role == 'hod':
            subj = get_object_or_404(Subject, id=subj_id, branch=faculty.department)
            sec = get_object_or_404(Section, id=sec_id, branch=faculty.department)
        else:
            subj = get_object_or_404(Subject, id=subj_id, faculty=faculty)
            sec = get_object_or_404(Section, id=sec_id)
            
        ex = get_object_or_404(Exam, id=ex_id)
        if ex.exam_type == 'final':
            messages.error(request, "Only the Administrator is authorized to upload Semester Final exam results.")
            return redirect(f"{request.path}?subject={subj_id}&exam={ex_id}&section={sec_id}")
            
        sec_students = Student.objects.filter(section=sec, is_active=True, user__is_deleted=False)
        
        # Check action type: CSV upload or Manual entry
        action = request.POST.get('action')
        
        if action == 'csv':
            if 'csv_file' not in request.FILES:
                messages.error(request, "Please upload a CSV file.")
                return redirect(f"{request.path}?subject={subj_id}&exam={ex_id}&section={sec_id}")
                
            csv_file = request.FILES['csv_file']
            if not csv_file.name.endswith('.csv'):
                messages.error(request, "Please upload a valid .csv file.")
                return redirect(f"{request.path}?subject={subj_id}&exam={ex_id}&section={sec_id}")
                
            try:
                data_set = csv_file.read().decode('utf-8-sig')
                io_string = io.StringIO(data_set)
                next(io_string, None) # skip header
                
                reader = csv.reader(io_string, delimiter=',', quotechar='"')
                success_count = 0
                errors = []
                
                with transaction.atomic():
                    for row_idx, row in enumerate(reader, start=2):
                        if not row or not row[0].strip():
                            continue
                        if len(row) < 2:
                            errors.append(f"Row {row_idx}: Missing columns.")
                            continue
                            
                        roll = row[0].strip().upper()
                        marks_str = row[1].strip()
                        max_str = row[2].strip() if len(row) > 2 and row[2].strip() else '100'
                        
                        try:
                            marks_obt = float(marks_str)
                            max_mks = float(max_str)
                        except ValueError:
                            errors.append(f"Row {row_idx}: Invalid marks format for {roll}.")
                            continue
                            
                        try:
                            # Verify student exists and belongs to the selected section
                            student = sec_students.get(roll_number=roll)
                            
                            Result.objects.update_or_create(
                                student=student,
                                exam=ex,
                                subject=subj,
                                defaults={
                                    'marks_obtained': marks_obt,
                                    'max_marks': max_mks,
                                    'grade': ''  # cleared so save() will auto-recompute
                                }
                            )
                            success_count += 1
                        except Student.DoesNotExist:
                            errors.append(f"Row {row_idx}: Student {roll} not found in section {sec}.")
                            
                if errors:
                    for err in errors[:5]:
                        messages.error(request, err)
                    if len(errors) > 5:
                        messages.error(request, f"...and {len(errors) - 5} more errors.")
                if success_count > 0:
                    messages.success(request, f"Successfully uploaded marks for {success_count} students in {sec}.")
                    
            except Exception as e:
                messages.error(request, f"Error uploading CSV: {e}")
                
        elif action == 'manual':
            try:
                success_count = 0
                max_marks_default = float(request.POST.get('max_marks_default', '100'))
                
                with transaction.atomic():
                    for stu in sec_students:
                        marks_input = request.POST.get(f"marks_{stu.id}", '').strip()
                        if marks_input == '':
                            continue
                            
                        try:
                            marks_obt = float(marks_input)
                        except ValueError:
                            messages.error(request, f"Invalid marks for student {stu.roll_number}.")
                            return redirect(f"{request.path}?subject={subj_id}&exam={ex_id}&section={sec_id}")
                            
                        Result.objects.update_or_create(
                            student=stu,
                            exam=ex,
                            subject=subj,
                            defaults={
                                'marks_obtained': marks_obt,
                                'max_marks': max_marks_default,
                                'grade': ''  # cleared so save() will auto-recompute
                            }
                        )
                        success_count += 1
                        
                messages.success(request, f"Successfully saved marks for {success_count} students in {sec}.")
            except Exception as e:
                messages.error(request, f"Error saving marks: {e}")
                
        if success_count > 0 and request.user.role == 'hod':
            try:
                from core.models import Notification
                Notification.objects.create(
                    title="HOD Marks Uploaded",
                    message=f"HOD {request.user.get_full_name() or request.user.username} uploaded marks for {subj.code} in section {sec.name} for exam {ex.name}.",
                    notif_type=Notification.TYPE_SYSTEM,
                    priority=Notification.PRIORITY_HIGH,
                    target_all=False,
                    target_role='admin',
                    created_by=request.user
                )
            except Exception as notif_err:
                logger.warning(f"Failed to create HOD mark upload notification: {notif_err}")

        return redirect(f"{request.path}?year={year_id}&subject={subj_id}&exam={ex_id}&section={sec_id}")
        
    context = {
        'years':               years,
        'year_id':             year_id,
        'subjects':            subjects,
        'exams':               exams,
        'selected_subject_id': selected_subject_id,
        'selected_exam_id':    selected_exam_id,
        'selected_section_id': selected_section_id,
        'selected_subject':    selected_subject,
        'selected_exam':       selected_exam,
        'selected_section':    selected_section,
        'sections':            sections,
        'students':            students,
        'current_results':     current_results,
    }
    return render(request, 'faculty/upload_marks.html', context)


@faculty_required
def add_achievement(request):
    faculty = request.faculty
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        category = request.POST.get('category', '').strip()
        date_str = request.POST.get('date_achieved', '').strip()

        if not (title and description and category and date_str):
            messages.error(request, "All fields are required.")
        else:
            try:
                date_achieved = datetime.date.fromisoformat(date_str)
                Achievement.objects.create(
                    user=request.user,
                    title=title,
                    description=description,
                    category=category,
                    date_achieved=date_achieved
                )
                messages.success(request, "Achievement submitted successfully. Pending HOD verification.")
                return redirect('faculty:add_achievement')
            except ValueError:
                messages.error(request, "Invalid date format.")
            except Exception as e:
                messages.error(request, f"Error saving achievement: {e}")

    achievements = Achievement.objects.filter(user=request.user).order_by('-date_achieved')
    return render(request, 'faculty/add_achievement.html', {
        'faculty': faculty,
        'achievements': achievements,
    })


# ─────────────────────────────────────────────
# FACULTY LEAVE REQUESTS
# ─────────────────────────────────────────────
@faculty_required
def leave_requests(request):
    """
    Faculty view to apply for a leave request and view leave history.
    Sends notifications to HOD & Admin upon submission.
    """
    faculty = request.faculty
    if request.method == 'POST':
        leave_type = request.POST.get('leave_type', 'casual').strip()
        start_date_str = request.POST.get('start_date', '').strip()
        end_date_str = request.POST.get('end_date', '').strip()
        reason = request.POST.get('reason', '').strip()
        substitute_notes = request.POST.get('substitute_notes', '').strip()

        if not (start_date_str and end_date_str and reason):
            messages.error(request, "Please provide Start Date, End Date, and Reason for leave.")
        else:
            try:
                start_date = datetime.date.fromisoformat(start_date_str)
                end_date = datetime.date.fromisoformat(end_date_str)

                if end_date < start_date:
                    messages.error(request, "End date cannot be prior to start date.")
                else:
                    leave_req = FacultyLeaveRequest.objects.create(
                        faculty=faculty,
                        leave_type=leave_type,
                        start_date=start_date,
                        end_date=end_date,
                        reason=reason,
                        substitute_notes=substitute_notes,
                        status='pending'
                    )

                    # Notify HOD & Admin via In-App Notification, Email & SMS
                    try:
                        from core.models import Notification
                        from accounts.models import User
                        from core.sms_utils import send_sms
                        from django.core.mail import send_mail
                        from django.conf import settings

                        dept_name = faculty.department.code if faculty.department else "General"
                        notif_title = f"Faculty Leave Request: {faculty.full_name}"
                        notif_msg = (
                            f"Leave Application from {faculty.full_name} ({faculty.employee_id}, {dept_name}): "
                            f"{leave_req.get_leave_type_display()} from {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}. "
                            f"Reason: {reason}"
                        )
                        
                        # 1. Notify Department HOD
                        if faculty.department:
                            hod_users = User.objects.filter(role='hod', faculty_profile__department=faculty.department)
                            for hod in hod_users:
                                # In-App Notification
                                Notification.objects.create(
                                    title=notif_title,
                                    message=notif_msg,
                                    notif_type=Notification.TYPE_ANNOUNCEMENT,
                                    priority=Notification.PRIORITY_HIGH,
                                    target_user=hod,
                                    target_role='hod',
                                    target_all=False,
                                    created_by=request.user
                                )
                                # Email Notification
                                if hod.email:
                                    send_mail(
                                        subject=f"[VVITU] {notif_title}",
                                        message=notif_msg,
                                        from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@vvitu.ac.in'),
                                        recipient_list=[hod.email],
                                        fail_silently=True
                                    )
                                # SMS Notification
                                if hasattr(hod, 'faculty_profile') and hod.faculty_profile.phone:
                                    send_sms(hod.faculty_profile.phone, f"VVITU: New leave request from {faculty.full_name} ({start_date.strftime('%d-%b')} to {end_date.strftime('%d-%b')}). Review on portal.")

                        # 2. Notify College Administration (Admins)
                        admin_users = User.objects.filter(role='admin')
                        for adm in admin_users:
                            # In-App Notification
                            Notification.objects.create(
                                title=notif_title,
                                message=notif_msg,
                                notif_type=Notification.TYPE_ANNOUNCEMENT,
                                priority=Notification.PRIORITY_HIGH,
                                target_user=adm,
                                target_role='admin',
                                target_all=False,
                                created_by=request.user
                            )
                            # Email Notification
                            if adm.email:
                                send_mail(
                                    subject=f"[VVITU] {notif_title}",
                                    message=notif_msg,
                                    from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@vvitu.ac.in'),
                                    recipient_list=[adm.email],
                                    fail_silently=True
                                )
                    except Exception as e:
                        logger.warning(f"Failed to notify admins for leave request: {e}")

                    messages.success(request, "Leave request submitted successfully! Pending approval from HOD or Admin.")
                    return redirect('faculty:leave_requests')
            except ValueError:
                messages.error(request, "Invalid date format.")
            except Exception as e:
                messages.error(request, f"Error submitting leave request: {e}")

    my_leaves = FacultyLeaveRequest.objects.filter(faculty=faculty).select_related('action_by').order_by('-created_at')
    
    total_leaves = my_leaves.count()
    approved_leaves = my_leaves.filter(status='approved').count()
    pending_leaves = my_leaves.filter(status='pending').count()
    rejected_leaves = my_leaves.filter(status='rejected').count()

    context = {
        'faculty': faculty,
        'my_leaves': my_leaves,
        'total_leaves': total_leaves,
        'approved_leaves': approved_leaves,
        'pending_leaves': pending_leaves,
        'rejected_leaves': rejected_leaves,
        'leave_type_choices': FacultyLeaveRequest.LEAVE_TYPE_CHOICES,
    }
    return render(request, 'faculty/leave_requests.html', context)


@faculty_required
def cancel_leave_request(request, pk):
    faculty = request.faculty
    leave_req = get_object_or_404(FacultyLeaveRequest, pk=pk, faculty=faculty)
    if leave_req.status == 'pending':
        leave_req.delete()
        messages.success(request, "Leave request cancelled successfully.")
    else:
        messages.error(request, "Cannot cancel a leave request that has already been actioned.")
    return redirect('faculty:leave_requests')


# ─────────────────────────────────────────────
# STUDENT COUNSELLING DOSSIER (FACULTY / COUNSELLOR)
# ─────────────────────────────────────────────
@faculty_required
def student_counselling_report(request, student_id):
    """
    View complete counselling dossier for a student assigned to or taught by this faculty member.
    """
    from core.counselling_utils import get_student_counselling_dossier
    from django.urls import reverse

    faculty = request.faculty
    student = get_object_or_404(Student, id=student_id, is_active=True, user__is_deleted=False)

    # Permission check: Is counsellor, class teacher, section instructor, or HOD/Admin?
    is_counsellor = (student.counsellor_id == faculty.id)
    is_class_teacher = (student.class_teacher_id == faculty.id)
    is_dept_hod = (request.user.role == 'hod' and faculty.department == student.branch)
    is_admin = (request.user.role == 'admin')
    
    # Check section instructor
    teaches_section = Timetable.objects.filter(faculty=faculty, section=student.section).exists() if student.section else False

    if not (is_counsellor or is_class_teacher or is_dept_hod or is_admin or teaches_section):
        messages.error(request, "You are not authorized to view this student's counselling report.")
        return redirect('faculty:counselled_students')

    dossier = get_student_counselling_dossier(student)
    context = {
        'dossier': dossier,
        'pdf_download_url': reverse('faculty:download_student_counselling_report_pdf', args=[student.id]),
        'back_url': reverse('faculty:counselled_students'),
    }
    return render(request, 'reports/counselling_report.html', context)


@faculty_required
def download_student_counselling_report_pdf(request, student_id):
    """
    Download official student counselling dossier PDF as faculty counsellor.
    """
    from core.counselling_utils import generate_counselling_report_pdf
    from django.http import HttpResponse

    faculty = request.faculty
    student = get_object_or_404(Student, id=student_id, is_active=True, user__is_deleted=False)

    # Permission check
    is_counsellor = (student.counsellor_id == faculty.id)
    is_class_teacher = (student.class_teacher_id == faculty.id)
    is_dept_hod = (request.user.role == 'hod' and faculty.department == student.branch)
    is_admin = (request.user.role == 'admin')
    teaches_section = Timetable.objects.filter(faculty=faculty, section=student.section).exists() if student.section else False

    if not (is_counsellor or is_class_teacher or is_dept_hod or is_admin or teaches_section):
        messages.error(request, "You are not authorized to download this student's counselling report.")
        return redirect('faculty:counselled_students')

    pdf_bytes = generate_counselling_report_pdf(student)
    filename = f"{student.roll_number}_Counselling_Dossier.pdf"
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


